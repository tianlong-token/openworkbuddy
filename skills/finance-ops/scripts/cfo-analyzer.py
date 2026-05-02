#!/usr/bin/env python3
"""
CFO Briefing Analyzer — Parse QuickBooks exports and generate executive financial summaries.
Outputs formatted briefing to stdout. Stores history for MoM comparison.

Usage:
    python3 cfo-analyzer.py --input ./data/uploads/
    python3 cfo-analyzer.py --input ./data/uploads/ --period 2025-01
    python3 cfo-analyzer.py --input ./data/uploads/ --no-history
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_dollar(val: Any) -> float:
    """Convert QB dollar string to float. Handles $, commas, parens for negatives, formula strings."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == '-' or s.lower() == 'none':
        return 0.0
    # Handle Excel formula strings like =-236705.50
    if s.startswith('='):
        s = s[1:]
        try:
            return float(s)
        except ValueError:
            return 0.0
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    s = s.replace('$', '').replace(',', '').replace('"', '').strip()
    if not s:
        return 0.0
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return 0.0


def status_emoji(value: float, green_range: tuple, yellow_range: tuple) -> str:
    """Return 🟢/🟡/🔴 based on thresholds. green_range/yellow_range are (min, max) inclusive."""
    gmin, gmax = green_range
    ymin, ymax = yellow_range
    if gmin <= value <= gmax:
        return "🟢"
    if ymin <= value <= ymax:
        return "🟡"
    return "🔴"


def pct(part: float, whole: float) -> float:
    return (part / whole * 100) if whole else 0.0


def fmt_k(val: float) -> str:
    """Format as $XXK or $X.XM."""
    if abs(val) >= 1_000_000:
        return f"${val/1_000_000:.2f}M"
    return f"${val/1_000:.0f}K"


def fmt_pct(val: float) -> str:
    return f"{val:.1f}%"


# ---------------------------------------------------------------------------
# File detection & parsing
# ---------------------------------------------------------------------------

def detect_file_type(filepath: Path) -> str | None:
    """Detect QB report type from file content."""
    ext = filepath.suffix.lower()

    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), data_only=False)
            for name in wb.sheetnames:
                nl = name.lower()
                if 'cash flow' in nl or 'statement of cash' in nl:
                    return 'cash_flow'
                if 'profit and loss detail' in nl or 'p&l detail' in nl:
                    return 'pl_detail'
                if 'profit and loss' in nl or 'p&l' in nl:
                    return 'pl_summary'
                if 'balance sheet' in nl:
                    return 'balance_sheet'
                if 'general ledger' in nl:
                    return 'general_ledger'
            ws = wb.active
            for row in ws.iter_rows(max_row=5, values_only=True):
                text = ' '.join(str(c).lower() for c in row if c)
                if 'statement of cash flow' in text:
                    return 'cash_flow'
                if 'profit and loss detail' in text:
                    return 'pl_detail'
                if 'profit and loss' in text:
                    return 'pl_summary'
                if 'balance sheet' in text:
                    return 'balance_sheet'
        except Exception:
            pass
        return None

    if ext != '.csv':
        return None

    try:
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            head = ''.join(f.readline() for _ in range(10)).lower()
    except Exception:
        return None

    if 'profit and loss by customer' in head or 'profit and loss by job' in head:
        return 'pl_by_customer'
    if 'profit and loss' in head:
        return 'pl_summary'
    if 'balance sheet' in head:
        return 'balance_sheet'
    if 'general ledger' in head:
        return 'general_ledger'
    if 'expenses by vendor' in head:
        return 'expenses_by_vendor'
    if 'transaction list by vendor' in head:
        return 'transactions_by_vendor'
    if 'bill payment' in head:
        return 'bill_payments'
    if 'account list' in head:
        return 'account_list'
    return None


def detect_period(filepath: Path) -> str | None:
    """Try to extract period from file header."""
    ext = filepath.suffix.lower()
    text = ''
    if ext == '.csv':
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                text = ''.join(f.readline() for _ in range(5))
        except Exception:
            pass
    elif ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), data_only=False)
            ws = wb.active
            for row in ws.iter_rows(max_row=5, values_only=True):
                text += ' '.join(str(c) for c in row if c) + '\n'
        except Exception:
            pass

    # Look for patterns like "February, 2025-January, 2026" or "March 2025"
    m = re.search(r'(\w+),?\s*(\d{4})\s*[-–]\s*(\w+),?\s*(\d{4})', text)
    if m:
        end_month, end_year = m.group(3), m.group(4)
        try:
            dt = datetime.strptime(f"{end_month} {end_year}", "%B %Y")
            return dt.strftime("%Y-%m")
        except ValueError:
            pass

    m = re.search(r'(\w+ \d{4})', text)
    if m:
        try:
            dt = datetime.strptime(m.group(1), "%B %Y")
            return dt.strftime("%Y-%m")
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# P&L Summary parser
# ---------------------------------------------------------------------------

def parse_pl_summary(filepath: Path) -> dict:
    """Parse P&L Summary CSV into structured data."""
    results: dict[str, Any] = {
        'revenue': {},
        'cogs': {},
        'expenses': {},
        'other_income': {},
        'other_expenses': {},
        'totals': {}
    }

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        lines = f.readlines()

    section = None

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Split on first comma that's not inside quotes
        parts = []
        in_quote = False
        current = ''
        for ch in line:
            if ch == '"':
                in_quote = not in_quote
            elif ch == ',' and not in_quote:
                parts.append(current.strip().strip('"'))
                current = ''
                continue
            current += ch
        parts.append(current.strip().strip('"'))

        account = parts[0] if parts else ''
        value = parse_dollar(parts[-1]) if len(parts) > 1 else 0.0

        # Track sections
        if account == 'Income':
            section = 'revenue'
            continue
        elif account == 'Cost of Goods Sold':
            section = 'cogs'
            continue
        elif account == 'Expenses':
            section = 'expenses'
            continue
        elif account == 'Other Income':
            section = 'other_income'
            continue
        elif account == 'Other Expenses':
            section = 'other_expenses'
            continue

        # Capture totals
        if account.startswith('Total for Income'):
            results['totals']['total_income'] = value
            section = None
        elif account.startswith('Total for Cost of Goods Sold'):
            results['totals']['total_cogs'] = value
        elif account.startswith('Gross Profit'):
            results['totals']['gross_profit'] = value
        elif account.startswith('Total for Expenses'):
            results['totals']['total_expenses'] = value
        elif account == 'Net Operating Income':
            results['totals']['net_operating_income'] = value
        elif account.startswith('Total for Other Income'):
            results['totals']['total_other_income'] = value
        elif account.startswith('Total for Other Expenses'):
            results['totals']['total_other_expenses'] = value
        elif account == 'Net Income':
            results['totals']['net_income'] = value
        elif account.startswith('Net Other Income'):
            results['totals']['net_other_income'] = value
        elif section and value != 0.0 and not account.startswith('Total for'):
            # Store individual line items
            clean_name = re.sub(r'^\d{5}\s+', '', account).strip()
            if section in results:
                results[section][clean_name] = value

    return results


# ---------------------------------------------------------------------------
# P&L by Customer parser
# ---------------------------------------------------------------------------

def parse_pl_by_customer(filepath: Path) -> dict:
    """Parse P&L by Customer CSV. Returns revenue by customer."""
    customers: dict[str, float] = {}

    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        lines = f.readlines()

    # Find header row with customer names
    header_idx = None
    headers = []
    for i, line in enumerate(lines):
        if 'Distribution account' in line:
            header_idx = i
            import csv
            reader = csv.reader([line])
            headers = next(reader)
            break

    if not headers:
        return {'customers': customers}

    # Find Total for Income row
    for line in lines[header_idx+1:]:
        if 'Total for Income' in line or 'Total for' in line and 'Revenue' in line:
            import csv
            reader = csv.reader([line])
            values = next(reader)
            for j, val in enumerate(values):
                if j > 0 and j < len(headers) and headers[j] and headers[j] != 'Total':
                    v = parse_dollar(val)
                    if v > 0:
                        name = headers[j].replace(' (deleted)', '').strip()
                        customers[name] = v
            break

    return {'customers': customers}


# ---------------------------------------------------------------------------
# Cash Flow XLSX parser
# ---------------------------------------------------------------------------

def parse_cash_flow(filepath: Path) -> dict:
    """Parse Cash Flow Statement XLSX."""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=False)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    result: dict[str, Any] = {'monthly_net_income': {}, 'monthly_net_cash': {}}

    # Find header row with month columns
    headers = []
    header_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(c) for c in row if c]
        for v in vals:
            if re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}', str(v)):
                headers = list(row)
                header_row_idx = i
                break
        if headers:
            break

    if not headers:
        return result

    for row in rows[header_row_idx+1:]:
        label = str(row[0]).strip() if row[0] else ''
        if 'Net Income' in label and 'reconcile' not in label.lower():
            for j, val in enumerate(row[1:], 1):
                if j < len(headers) and headers[j]:
                    month_str = str(headers[j])
                    v = parse_dollar(val)
                    result['monthly_net_income'][month_str] = v
        if 'net cash' in label.lower() or 'net change in cash' in label.lower():
            for j, val in enumerate(row[1:], 1):
                if j < len(headers) and headers[j]:
                    month_str = str(headers[j])
                    v = parse_dollar(val)
                    result['monthly_net_cash'][month_str] = v

    return result


# ---------------------------------------------------------------------------
# KPI computation
# ---------------------------------------------------------------------------

def compute_kpis(pl_data: dict, customer_data: dict | None, cash_flow_data: dict | None) -> dict:
    """Compute all KPIs from parsed data."""
    kpis: dict[str, Any] = {}
    totals = pl_data.get('totals', {})
    revenue_items = pl_data.get('revenue', {})
    cogs_items = pl_data.get('cogs', {})
    expense_items = pl_data.get('expenses', {})
    other_exp = pl_data.get('other_expenses', {})
    other_inc = pl_data.get('other_income', {})

    # --- Revenue ---
    total_revenue = totals.get('total_income', 0)
    kpis['total_revenue'] = total_revenue

    # Revenue by service line
    service_revenue = {}
    for k, v in revenue_items.items():
        if v != 0:
            service_revenue[k] = v
    kpis['revenue_by_service'] = dict(sorted(service_revenue.items(), key=lambda x: -x[1]))

    # --- COGS & Gross Margin ---
    total_cogs = totals.get('total_cogs', 0)
    gross_profit = totals.get('gross_profit', total_revenue - total_cogs)
    kpis['total_cogs'] = total_cogs
    kpis['gross_profit'] = gross_profit
    kpis['gross_margin_pct'] = pct(gross_profit, total_revenue)

    # --- Net Income ---
    kpis['net_income'] = totals.get('net_income', 0)
    kpis['net_operating_income'] = totals.get('net_operating_income', 0)
    kpis['net_margin_pct'] = pct(kpis['net_income'], total_revenue)

    # --- People Costs ---
    salary_total = 0
    contractor_total = 0
    payroll_tax_total = 0
    benefits_total = 0

    all_items = {}
    all_items.update(cogs_items)
    all_items.update(expense_items)

    for k, v in all_items.items():
        kl = k.lower()
        if 'salari' in kl or 'wages' in kl or 'payroll -' in kl or 'payroll - ' in kl:
            salary_total += v
        elif 'contractor' in kl:
            contractor_total += v
        elif 'payroll tax' in kl:
            payroll_tax_total += v
        elif 'benefit' in kl or 'insurance' in kl and 'benefit' in kl:
            benefits_total += v
        elif 'commission' in kl:
            contractor_total += v

    people_total = salary_total + contractor_total + payroll_tax_total + benefits_total
    kpis['people_costs'] = {
        'total': people_total,
        'salaries': salary_total,
        'contractors': contractor_total,
        'payroll_taxes': payroll_tax_total,
        'benefits': benefits_total,
        'pct_of_revenue': pct(people_total, total_revenue),
        'contractor_pct': pct(contractor_total, people_total) if people_total else 0
    }

    # --- Tool/Subscription Costs ---
    tool_total = 0
    tool_items = {}
    for k, v in all_items.items():
        kl = k.lower()
        if any(w in kl for w in ['subscription', 'tools', 'hosting', 'it hosting', 'software']):
            tool_total += v
            tool_items[k] = v

    kpis['tool_costs'] = {
        'total': tool_total,
        'pct_of_revenue': pct(tool_total, total_revenue),
        'items': dict(sorted(tool_items.items(), key=lambda x: -x[1]))
    }

    # --- Expense Categories ---
    total_expenses = totals.get('total_expenses', 0)
    kpis['total_opex'] = total_expenses

    categories = {
        'Sales & Growth': 0,
        'Marketing & Branding': 0,
        'G&A': 0,
        'Facilities': 0,
        'Other Expenses': 0,
    }

    for k, v in expense_items.items():
        kl = k.lower()
        if any(w in kl for w in ['sales', 'commission', 'growth']):
            categories['Sales & Growth'] += v
        elif any(w in kl for w in ['marketing', 'advertising', 'writing', 'content', 'podcast', 'branding', 'networking']):
            categories['Marketing & Branding'] += v
        elif any(w in kl for w in ['rent', 'facilit']):
            categories['Facilities'] += v
        else:
            categories['G&A'] += v

    kpis['expense_categories'] = {k: {'amount': v, 'pct_of_revenue': pct(v, total_revenue)} for k, v in categories.items() if v}

    # --- Interest & Debt ---
    interest = other_exp.get('Interest Expense', 0)
    amortization = other_exp.get('Amortization Expense', 0)
    kpis['interest_expense'] = interest
    kpis['amortization'] = amortization
    kpis['interest_pct_of_revenue'] = pct(interest, total_revenue)

    # --- Other Income ---
    kpis['other_income'] = {k: v for k, v in other_inc.items() if v}
    kpis['total_other_income'] = totals.get('total_other_income', 0)

    # --- Notable Expense Items (anomaly detection) ---
    notable = {}
    for k, v in all_items.items():
        kl = k.lower()
        if v > 5000 and not any(w in kl for w in ['salari', 'wages', 'payroll', 'rent', 'amortization']):
            notable[k] = v
    kpis['notable_expenses'] = dict(sorted(notable.items(), key=lambda x: -x[1])[:15])

    # --- Recruiting ---
    recruiting = 0
    for k, v in all_items.items():
        if 'recruit' in k.lower():
            recruiting += v
    kpis['recruiting_spend'] = recruiting

    # --- Owner Expenses ---
    owner_total = 0
    for k, v in all_items.items():
        if 'owner' in k.lower() or k.startswith('OE -'):
            owner_total += v
    kpis['owner_expenses'] = owner_total

    # --- Customer Data ---
    if customer_data and customer_data.get('customers'):
        custs = customer_data['customers']
        sorted_custs = sorted(custs.items(), key=lambda x: -x[1])
        kpis['top_customers'] = sorted_custs[:10]
        kpis['customer_count'] = len([c for c in custs.values() if c > 0])
        if sorted_custs and total_revenue:
            top_pct = sorted_custs[0][1] / total_revenue * 100
            kpis['top_customer_concentration'] = (sorted_custs[0][0], top_pct)

    # --- Cash Flow Data ---
    if cash_flow_data:
        monthly_ni = cash_flow_data.get('monthly_net_income', {})
        if monthly_ni:
            vals = list(monthly_ni.values())
            kpis['monthly_net_income'] = monthly_ni
            if len(vals) >= 3:
                kpis['last_3mo_avg_ni'] = sum(vals[-3:]) / 3
            kpis['monthly_burn'] = abs(min(vals)) if any(v < 0 for v in vals) else 0

    return kpis


# ---------------------------------------------------------------------------
# MoM comparison
# ---------------------------------------------------------------------------

def load_prior_period(history_dir: Path, current_period: str) -> dict | None:
    """Load most recent prior period from history."""
    if not history_dir.exists():
        return None

    files = sorted(history_dir.glob('*.json'), reverse=True)
    for f in files:
        if f.stem != current_period:
            try:
                with open(f) as fh:
                    return json.load(fh)
            except Exception:
                continue
    return None


def compute_variance(current: float, prior: float) -> str:
    """Format MoM variance."""
    if prior == 0:
        return "N/A"
    change = ((current - prior) / abs(prior)) * 100
    arrow = "↑" if change > 0 else "↓" if change < 0 else "→"
    return f"{arrow} {abs(change):.1f}%"


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def format_briefing(kpis: dict, prior: dict | None, period: str) -> str:
    """Format KPIs into executive briefing with status indicators."""
    lines = []
    lines.append(f"*📊 CFO Briefing — {period}*")
    lines.append("=" * 40)

    rev = kpis['total_revenue']

    # --- Revenue ---
    lines.append("")
    lines.append("*💰 Revenue*")
    lines.append(f"• Total Revenue: *{fmt_k(rev)}*" + (f" ({compute_variance(rev, prior.get('total_revenue', 0))} MoM)" if prior else ""))

    if kpis.get('revenue_by_service'):
        top3 = list(kpis['revenue_by_service'].items())[:5]
        for name, val in top3:
            lines.append(f"  · {name}: {fmt_k(val)} ({fmt_pct(pct(val, rev))})")

    # --- Profitability ---
    gm = kpis['gross_margin_pct']
    nm = kpis['net_margin_pct']
    gm_status = status_emoji(gm, (60, 999), (45, 60))
    nm_status = status_emoji(nm, (10, 999), (0, 10))

    lines.append("")
    lines.append("*📈 Profitability*")
    lines.append(f"• {gm_status} Gross Margin: *{fmt_pct(gm)}* (Gross Profit: {fmt_k(kpis['gross_profit'])})")
    lines.append(f"• {nm_status} Net Income: *{fmt_k(kpis['net_income'])}* ({fmt_pct(nm)} margin)")
    lines.append(f"• Net Operating Income: {fmt_k(kpis['net_operating_income'])}")

    if prior:
        lines.append(f"  MoM: Revenue {compute_variance(rev, prior.get('total_revenue', 0))}, "
                     f"Net Income {compute_variance(kpis['net_income'], prior.get('net_income', 0))}")

    # --- People Costs ---
    pc = kpis['people_costs']
    pc_status = status_emoji(pc['pct_of_revenue'], (0, 65), (65, 75))
    contr_status = status_emoji(pc['contractor_pct'], (0, 30), (30, 50))

    lines.append("")
    lines.append("*👥 People Costs*")
    lines.append(f"• {pc_status} Total: *{fmt_k(pc['total'])}* ({fmt_pct(pc['pct_of_revenue'])} of revenue)")
    lines.append(f"  · Salaries: {fmt_k(pc['salaries'])}")
    lines.append(f"  · {contr_status} Contractors: {fmt_k(pc['contractors'])} ({fmt_pct(pc['contractor_pct'])} of people costs)")
    lines.append(f"  · Payroll Taxes: {fmt_k(pc['payroll_taxes'])}")
    lines.append(f"  · Benefits: {fmt_k(pc['benefits'])}")

    # --- Tools & Subscriptions ---
    tc = kpis['tool_costs']
    tc_status = status_emoji(tc['pct_of_revenue'], (0, 8), (8, 12))

    lines.append("")
    lines.append("*🔧 Tools & Subscriptions*")
    lines.append(f"• {tc_status} Total: *{fmt_k(tc['total'])}* ({fmt_pct(tc['pct_of_revenue'])} of revenue)")
    if tc['items']:
        for name, val in list(tc['items'].items())[:5]:
            lines.append(f"  · {name}: {fmt_k(val)}")

    # --- Expense Categories ---
    if kpis.get('expense_categories'):
        lines.append("")
        lines.append("*📋 Operating Expenses*")
        lines.append(f"• Total OpEx: *{fmt_k(kpis['total_opex'])}* ({fmt_pct(pct(kpis['total_opex'], rev))} of revenue)")
        for cat, data in kpis['expense_categories'].items():
            lines.append(f"  · {cat}: {fmt_k(data['amount'])} ({fmt_pct(data['pct_of_revenue'])})")

    # --- Interest & Debt ---
    int_status = status_emoji(kpis['interest_pct_of_revenue'], (0, 3), (3, 5))
    lines.append("")
    lines.append("*🏦 Debt & Interest*")
    lines.append(f"• {int_status} Interest Expense: *{fmt_k(kpis['interest_expense'])}* ({fmt_pct(kpis['interest_pct_of_revenue'])} of revenue)")
    lines.append(f"• Amortization (non-cash): {fmt_k(kpis['amortization'])}")

    # --- Notable Items ---
    if kpis.get('recruiting_spend'):
        lines.append(f"• Recruiting: {fmt_k(kpis['recruiting_spend'])}")

    if kpis.get('owner_expenses'):
        lines.append(f"• Owner Expenses: {fmt_k(kpis['owner_expenses'])}")

    # --- Customer Concentration ---
    if kpis.get('top_customers'):
        lines.append("")
        lines.append("*🏢 Top Customers*")
        for name, val in kpis['top_customers'][:5]:
            conc = pct(val, rev)
            conc_flag = " ⚠️" if conc > 15 else ""
            lines.append(f"  · {name}: {fmt_k(val)} ({fmt_pct(conc)}){conc_flag}")
        if kpis.get('top_customer_concentration'):
            cname, cpct = kpis['top_customer_concentration']
            conc_status = status_emoji(100 - cpct, (75, 100), (60, 75))
            lines.append(f"• {conc_status} Top client concentration: {cname} at {fmt_pct(cpct)}")
        if kpis.get('customer_count'):
            lines.append(f"• Active customers: {kpis['customer_count']}")

    # --- Other Income ---
    if kpis.get('other_income'):
        lines.append("")
        lines.append("*📥 Other Income*")
        for k, v in kpis['other_income'].items():
            lines.append(f"  · {k}: {fmt_k(v)}")

    # --- Monthly Trends ---
    if kpis.get('monthly_net_income'):
        lines.append("")
        lines.append("*📉 Monthly Net Income Trend*")
        for month, val in kpis['monthly_net_income'].items():
            ml = month.lower()
            if 'total' in ml:
                continue
            if val == 0.0:
                continue
            indicator = "✅" if val > 0 else "❌"
            lines.append(f"  · {month}: {fmt_k(val)} {indicator}")

    # --- Alerts ---
    alerts = []
    if kpis['gross_margin_pct'] < 45:
        alerts.append("🔴 Gross margin critically low (<45%). Target: 60%+")
    if pc['pct_of_revenue'] > 75:
        alerts.append("🔴 People costs >75% of revenue. Target: 55-65%")
    if pc['contractor_pct'] > 50:
        alerts.append("🟡 Contractor spend >50% of people costs — dependency risk")
    if kpis['interest_pct_of_revenue'] > 5:
        alerts.append("🔴 Interest >5% of revenue — debt load is heavy")
    if kpis.get('recruiting_spend', 0) > 80000:
        alerts.append("🟡 Recruiting spend elevated — review if active hires justify")
    if kpis.get('owner_expenses', 0) > 100000:
        alerts.append("🟡 Owner expenses >$100K TTM")
    if kpis['net_income'] < 0:
        deficit = abs(kpis['net_income'])
        alerts.append(f"🔴 Operating at a loss: {fmt_k(deficit)} deficit")

    if alerts:
        lines.append("")
        lines.append("*⚠️ Alerts*")
        for a in alerts:
            lines.append(f"• {a}")

    lines.append("")
    lines.append("_Data from QuickBooks exports. Review with your finance team before acting._")

    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='CFO Briefing Analyzer')
    parser.add_argument('--input', '-i', default='./data/uploads/',
                       help='Directory containing QB export files')
    parser.add_argument('--period', '-p', help='Override period label (YYYY-MM)')
    parser.add_argument('--history', default='./data/history/',
                       help='History directory for MoM comparison')
    parser.add_argument('--no-history', action='store_true', help='Skip saving to history')
    args = parser.parse_args()

    input_dir = Path(args.input)
    history_dir = Path(args.history)

    if not input_dir.exists():
        print(f"Error: Input directory not found: {input_dir}", file=sys.stderr)
        sys.exit(1)

    # Find and classify files
    files = list(input_dir.glob('*.csv')) + list(input_dir.glob('*.xlsx')) + list(input_dir.glob('*.xls'))
    if not files:
        print(f"Error: No CSV/XLSX files found in {input_dir}", file=sys.stderr)
        sys.exit(1)

    classified: dict[str, Path] = {}
    period = args.period

    for f in files:
        ftype = detect_file_type(f)
        if ftype:
            classified[ftype] = f
            if not period:
                p = detect_period(f)
                if p:
                    period = p
            print(f"  Detected: {f.name} → {ftype}", file=sys.stderr)
        else:
            print(f"  Skipped (unknown format): {f.name}", file=sys.stderr)

    if not period:
        period = datetime.now().strftime("%Y-%m")

    print(f"  Period: {period}", file=sys.stderr)
    print(f"  Files classified: {len(classified)}", file=sys.stderr)

    # Parse files
    pl_data = {}
    customer_data = None
    cash_flow_data = None

    if 'pl_summary' in classified:
        pl_data = parse_pl_summary(classified['pl_summary'])
    elif 'pl_by_customer' not in classified:
        print("Warning: No P&L file found. Output will be limited.", file=sys.stderr)
        pl_data = {'revenue': {}, 'cogs': {}, 'expenses': {}, 'other_income': {}, 'other_expenses': {}, 'totals': {}}

    if 'pl_by_customer' in classified:
        customer_data = parse_pl_by_customer(classified['pl_by_customer'])
        if not pl_data.get('totals'):
            pl_data = parse_pl_summary(classified['pl_by_customer'])

    if 'cash_flow' in classified:
        cash_flow_data = parse_cash_flow(classified['cash_flow'])

    # Compute KPIs
    kpis = compute_kpis(pl_data, customer_data, cash_flow_data)

    # Load prior period
    prior = None
    if history_dir.exists():
        prior = load_prior_period(history_dir, period)

    # Save current period
    if not args.no_history:
        history_dir.mkdir(parents=True, exist_ok=True)
        history_file = history_dir / f"{period}.json"

        save_data = {
            'period': period,
            'total_revenue': kpis['total_revenue'],
            'gross_profit': kpis['gross_profit'],
            'gross_margin_pct': kpis['gross_margin_pct'],
            'net_income': kpis['net_income'],
            'net_margin_pct': kpis['net_margin_pct'],
            'total_cogs': kpis['total_cogs'],
            'total_opex': kpis['total_opex'],
            'people_costs_total': kpis['people_costs']['total'],
            'people_costs_pct': kpis['people_costs']['pct_of_revenue'],
            'tool_costs_total': kpis['tool_costs']['total'],
            'tool_costs_pct': kpis['tool_costs']['pct_of_revenue'],
            'interest_expense': kpis['interest_expense'],
            'recruiting_spend': kpis.get('recruiting_spend', 0),
            'owner_expenses': kpis.get('owner_expenses', 0),
            'customer_count': kpis.get('customer_count', 0),
            'timestamp': datetime.now().isoformat(),
        }
        with open(history_file, 'w') as f:
            json.dump(save_data, f, indent=2)
        print(f"  Saved history: {history_file}", file=sys.stderr)

    # Output briefing
    briefing = format_briefing(kpis, prior, period)
    print(briefing)


if __name__ == '__main__':
    main()
